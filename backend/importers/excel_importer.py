"""
Excel 导入器

从 Excel 课表文件导入课程数据
支持多种格式：
1. 强智系统导出的 Excel 格式
2. 标准课表格式（自动检测表头）
3. 简化格式（课程名 教师 地点）
"""

import uuid
import re
from typing import List, Tuple, Optional, Dict
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

try:
    from .base_importer import BaseImporter
    from .qiangzhi_importer import QiangZhiImporter
    from ..models.course_base import CourseBase
    from ..models.course_detail import CourseDetail
    from ..models.week_type import WeekType
    from ..utils.color_manager import ColorManager
except ImportError:
    from importers.base_importer import BaseImporter
    from importers.qiangzhi_importer import QiangZhiImporter
    from models.course_base import CourseBase
    from models.course_detail import CourseDetail
    from models.week_type import WeekType
    from utils.color_manager import ColorManager


class ExcelImporter(BaseImporter):
    """
    Excel 导入器
    
    智能解析 Excel 课表文件
    - 自动检测表头位置和格式
    - 支持强智系统导出格式
    - 支持标准课表格式
    """
    
    def __init__(self):
        """初始化 Excel 导入器"""
        self.color_manager = ColorManager()
        self.qiangzhi_importer = None  # 延迟初始化
        
        # 周次模式：{第2-16周 或 {第2-16周(单) 或 {第2-16周(双)
        self.week_pattern = re.compile(r'\{第(\d{1,2})[-]*(\d*)周(?:\((单|双)\))?')
        
        # 🔥 新增：支持 [周] 格式的周次模式，如 "2-9,11-17[周]"
        self.bracket_week_pattern = re.compile(r'([\d,\-]+)\[周\]')
        
        # 🔥 新增：节次模式，匹配 [01-02-03-04节] 或 [1-2节] 等格式
        self.section_pattern = re.compile(r'\[(\d+(?:-\d+)*)\s*节?\]')
        
        # 用于模糊匹配表头的关键字
        self.weekday_map = {
            "一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "日": 7, "天": 7
        }
        
        # 星期映射（完整版）
        self.weekday_names = {
            "周一": 1, "星期一": 1, "Monday": 1, "Mon": 1,
            "周二": 2, "星期二": 2, "Tuesday": 2, "Tue": 2,
            "周三": 3, "星期三": 3, "Wednesday": 3, "Wed": 3,
            "周四": 4, "星期四": 4, "Thursday": 4, "Thu": 4,
            "周五": 5, "星期五": 5, "Friday": 5, "Fri": 5,
            "周六": 6, "星期六": 6, "Saturday": 6, "Sat": 6,
            "周日": 7, "星期日": 7, "Sunday": 7, "Sun": 7,
        }
    
    def get_supported_formats(self) -> List[str]:
        """获取支持的文件格式"""
        return ['.xlsx', '.xls']
    
    def validate(self, content: str) -> Tuple[bool, str]:
        """
        验证 Excel 文件路径是否有效
        """
        if not content or not content.strip():
            return False, "文件路径为空"
        
        file_path = content
        if file_path.endswith('.xlsx'):
            try:
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                workbook.close()
                return True, ""
            except Exception as e:
                return False, f"无法打开 XLSX 文件: {str(e)}"
        elif file_path.endswith('.xls'):
            try:
                import xlrd
                workbook = xlrd.open_workbook(file_path)
                return True, ""
            except ImportError:
                return False, "由于缺少 'xlrd' 库，无法打开 .xls 文件。请安装: pip install xlrd"
            except Exception as e:
                return False, f"无法打开 XLS 文件: {str(e)}"
        
        return False, "不支持的文件格式"
    
    def parse(self, file_path: str) -> Tuple[List[CourseBase], List[CourseDetail]]:
        """
        解析 Excel 文件
        """
        # 验证文件
        valid, msg = self.validate(file_path)
        if not valid:
            raise ValueError(msg)
        
        if file_path.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            # 这里的 sheet 已经有 iter_rows
        else:
            import xlrd
            xl_workbook = xlrd.open_workbook(file_path)
            xl_sheet = xl_workbook.sheet_by_index(0)
            
            # 创建一个兼容 openpyxl iter_rows 的包装对象
            class SheetWrapper:
                def __init__(self, sheet):
                    self.sheet = sheet
                def iter_rows(self, values_only=True, max_row=None):
                    rows = []
                    num_rows = self.sheet.nrows
                    if max_row: num_rows = min(num_rows, max_row)
                    for r in range(num_rows):
                        row_values = [self.sheet.cell_value(r, c) for c in range(self.sheet.ncols)]
                        rows.append(row_values)
                    return rows
            sheet = SheetWrapper(xl_sheet)

        try:
            # 检测格式类型
            format_type = self._detect_format(sheet)
            
            # 根据格式类型选择解析方法
            if format_type == "qiangzhi":
                import_beans = self._parse_qiangzhi_format(sheet)
            else:
                import_beans = self._parse_standard_format(sheet)
            
            # 转换为 CourseBase 和 CourseDetail
            course_bases, course_details = self._convert_to_courses(import_beans)
            
            return course_bases, course_details
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise ValueError(f"解析 Excel 文件失败: {str(e)}")
    
    def _detect_format(self, sheet: Worksheet) -> str:
        """
        检测 Excel 格式类型
        
        Args:
            sheet: openpyxl 工作表对象
            
        Returns:
            格式类型：'qiangzhi' 或 'standard'
        """
        # 检查前几行是否包含强智系统特征
        rows = list(sheet.iter_rows(values_only=True, max_row=10))
        
        for row in rows:
            for cell in row:
                if cell and isinstance(cell, str):
                    # 🔥 A. 扩展格式识别：支持 {第...周} 和 [周] 两种格式
                    # 检查是否包含强智系统特征：{第...周} 或 [周]
                    has_week_marker = ('{第' in cell and '周' in cell) or '[周]' in cell
                    # 同时检查是否有节次标记 [01 或 [1
                    has_section_marker = '[0' in cell or '[1' in cell
                    
                    if has_week_marker and has_section_marker:
                        print(f"✅ [Excel] 检测到强智/结构化课表格式: 包含周次标记和节次标记")
                        return "qiangzhi"
        
        return "standard"
    
    def _parse_qiangzhi_format(self, sheet: Worksheet) -> List[dict]:
        """
        解析强智系统格式的 Excel - 增强版：支持自动定位和多行内容拆分
        
        Args:
            sheet: openpyxl 工作表对象
            
        Returns:
            课程信息字典列表
        """
        import_beans = []
        
        # 获取所有行
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return import_beans
        
        # 智能查找表头行和列映射
        header_row_idx, weekday_col_map = self._find_header_and_columns(sheet)
        
        # 查找节次列（第一列通常是节次）
        section_col_idx = 0
        
        # 遍历数据行
        for row_idx in range(header_row_idx + 1, len(rows)):
            row = rows[row_idx]
            if not row: continue
            
            # 提取节次信息
            current_section = self._extract_section_num(row[section_col_idx])
            if current_section is None:
                # 如果第一列没数字，尝试从行号推断
                current_section = ((row_idx - header_row_idx - 1) * 2) + 1
            
            # 遍历所有被识别为星期列的列
            for col_idx, day_of_week in weekday_col_map.items():
                if col_idx >= len(row): continue
                cell_value = row[col_idx]
                
                if not cell_value or str(cell_value).strip() == "":
                    continue
                
                cell_text = str(cell_value)
                
                # 直接交给 _parse_qiangzhi_cell 处理，它内部有更完善的智能拆分逻辑
                beans = self._parse_qiangzhi_cell(cell_text, current_section, day_of_week)
                import_beans.extend(beans)
        
        return import_beans
    
    def _find_header_and_columns(self, sheet: Worksheet) -> Tuple[int, Dict[int, int]]:
        """
        自动搜索包含"星期"或"周"的行，并识别各列对应的星期
        
        Args:
            sheet: openpyxl 工作表对象
            
        Returns:
            (表头行索引, 列索引->星期数字的映射)
        """
        # 获取所有行以便处理（如果是 xlrd 读入的，已经是 values_only 的列表）
        rows = list(sheet.iter_rows(values_only=True, max_row=10))
        
        for row_idx, row in enumerate(rows):
            col_map = {}
            found_weekday = False
            for col_idx, cell in enumerate(row):
                if not cell:
                    continue
                cell_str = str(cell).strip()
                
                # 精确匹配和模糊匹配结合
                for weekday_name, weekday_num in self.weekday_names.items():
                    if weekday_name == cell_str or weekday_name in cell_str:
                        col_map[col_idx] = weekday_num
                        found_weekday = True
                        break
            
            if found_weekday:
                return row_idx, col_map
        
        # 默认回退：假设第1列是节次，第2列=周一 ... 第8列=周日
        return 0, {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 7}


    def _parse_qiangzhi_cell(self, content: str, section: int, day: int) -> List[dict]:
        """
        解析强智系统单元格内容 (增强版)
        
        增强功能:
        - 支持 {第...周} 和 [周] 两种周次格式
        - 支持离散周段（如 2-9,11-17[周]）
        - 支持多门课程紧密连接（无换行符）
        - 优先从单元格提取节次，其次使用行头
        - 防止同一门课在多行重复导入
        - 添加详细的解析日志
        
        Args:
            content: 单元格文本内容
            section: 传入的节次（从行头获取，作为默认值）
            day: 星期几（1-7）
            
        Returns:
            课程信息字典列表
        """
        courses = []
        # 清理内容
        content = content.replace('\r', '').strip()
        if not content: return courses

        # 🔥 检查是否包含周次标记
        week_match = self.week_pattern.search(content)
        bracket_week_match = self.bracket_week_pattern.search(content)
        
        if not week_match and not bracket_week_match:
            print(f"⚠️ [Excel] 未找到周次标记，跳过单元格")
            return courses
        
        # 🔥 新增：智能拆分多门课程
        # 检测是否有多个周次标记（表示多门课程）
        week_marker_count = len(self.week_pattern.findall(content)) + len(self.bracket_week_pattern.findall(content))
        
        if week_marker_count > 1:
            # 多门课程：按周次标记拆分
            course_segments = self._split_multiple_courses(content)
            print(f"📝 [Excel] 检测到 {len(course_segments)} 门课程在同一单元格")
            
            for seg in course_segments:
                seg_courses = self._parse_single_course(seg, section, day)
                courses.extend(seg_courses)
        else:
            # 单门课程：直接解析
            seg_courses = self._parse_single_course(content, section, day)
            courses.extend(seg_courses)
        
        return courses
    
    def _split_multiple_courses(self, content: str) -> List[str]:
        """
        拆分包含多门课程的单元格内容
        
        策略：
        1. 找到所有周次标记的位置
        2. 向前查找课程名称的起始位置（通常在地点信息之后）
        3. 按课程边界拆分
        
        Args:
            content: 单元格内容
            
        Returns:
            课程片段列表
        """
        segments = []
        
        # 找到所有周次标记的位置
        week_positions = []
        
        # 查找 {第...周} 格式
        for match in self.week_pattern.finditer(content):
            week_positions.append(('brace', match.start(), match.end()))
        
        # 查找 [周] 格式
        for match in self.bracket_week_pattern.finditer(content):
            week_positions.append(('bracket', match.start(), match.end()))
        
        # 按位置排序
        week_positions.sort(key=lambda x: x[1])
        
        if not week_positions:
            return [content]
        
        # 🔥 改进：找到每门课程的起始位置
        # 策略：地点信息结束后就是下一门课的开始
        # 地点格式：【校区】楼栋+房间号（如：【红湘校区】南华楼202）
        course_starts = [0]  # 第一门课从开头开始
        
        # 查找所有地点信息的结束位置
        # 地点模式：【...】...数字
        location_pattern = re.compile(r'[【\[]([^】\]]+)[】\]]([^\【\[]*?\d+)')
        
        for match in location_pattern.finditer(content):
            end_pos = match.end()
            # 检查这个地点后面是否紧跟着新的课程（不是当前周次标记的一部分）
            # 如果后面有周次标记，说明这是一个课程边界
            for i, (marker_type, week_start, week_end) in enumerate(week_positions):
                if i > 0 and end_pos < week_start and end_pos > week_positions[i-1][2]:
                    # 这个地点在两个周次标记之间，是课程边界
                    if end_pos not in course_starts:
                        course_starts.append(end_pos)
                    break
        
        # 排序并去重
        course_starts = sorted(set(course_starts))
        
        # 根据起始位置拆分
        for i in range(len(course_starts)):
            if i < len(course_starts) - 1:
                segment = content[course_starts[i]:course_starts[i+1]].strip()
            else:
                segment = content[course_starts[i]:].strip()
            
            if segment:
                segments.append(segment)
        
        return segments
    
    def _parse_single_course(self, content: str, section: int, day: int) -> List[dict]:
        """
        解析单门课程的内容
        
        Args:
            content: 单门课程的文本内容
            section: 传入的节次
            day: 星期几
            
        Returns:
            课程信息字典列表（可能有多个周段）
        """
        courses = []
        
        # 🔥 C. 节次优先从单元格提取，其次使用行头
        section_match = self.section_pattern.search(content)
        actual_section = section  # 默认使用传入的section（行头）
        duration = 2  # 默认2节课
        
        if section_match:
            # 提取节次字符串，如 "01-02-03-04" 或 "1-2"
            section_str = section_match.group(1)
            # 提取所有数字
            section_numbers = [int(n) for n in re.findall(r'\d+', section_str)]
            
            if section_numbers:
                actual_section = min(section_numbers)
                duration = max(section_numbers) - min(section_numbers) + 1
                print(f"✅ [Excel] 节次解析成功: '{section_str}' → start={actual_section}, duration={duration}")
                
                # When a cell contains an explicit section marker, trust it.
                if section != actual_section:
                    # Keep dedup protection for long merged blocks (e.g. 01-02-03-04),
                    # but allow short blocks like [1-2节] to avoid false negatives.
                    if duration >= 3:
                        print(f"⚠️ [Excel] 行头节次({section}) 与单元格节次({actual_section})不一致，判定为跨行重复，跳过")
                        return []
                    print(f"ℹ️ [Excel] 行头节次({section}) 与单元格节次({actual_section})不一致，按单元格节次导入")
            else:
                print(f"⚠️ [Excel] 节次解析失败: 无法从 '{section_str}' 提取数字，使用行头 section={section}")
        else:
            print(f"📝 [Excel] 未找到节次标记，使用行头节次: section={section}, duration=2")
        
        # 🔥 提取课程信息：课程名、教师、地点
        name = "未知课程"
        teacher = ""
        location = ""
        
        # 找到周次标记的位置
        week_match = self.week_pattern.search(content)
        bracket_week_match = self.bracket_week_pattern.search(content)
        
        week_start_pos = -1
        if bracket_week_match:
            week_start_pos = bracket_week_match.start()
        elif week_match:
            week_start_pos = week_match.start()
        
        if week_start_pos > 0:
            # 周次标记之前的内容包含：课程名 + 教师
            before_week = content[:week_start_pos].strip()
            
            # 🔥 增强版提取：处理带换行的情况
            lines = [l.strip() for l in before_week.split('\n') if l.strip()]
            
            if lines:
                # 策略：检查最后一行是否匹配 "教师名()"
                last_line = lines[-1]
                paren_match = re.search(r'([^\n()（）]+(?:\([^)]*\)|（[^）]*）))$', last_line)
                
                if paren_match:
                    # 匹配到 "教师名()"
                    teacher = paren_match.group(1).strip()
                    # 如果这行除了教师名还有别的内容（如：课程名教师名()），则拆分
                    name_part = last_line[:paren_match.start()].strip()
                    if name_part:
                        name = "\n".join(lines[:-1] + [name_part]).strip()
                    else:
                        name = "\n".join(lines[:-1]).strip()
                        if not name and len(lines) == 1 and teacher:
                            teacher_suffix_match = re.search(r'(\([^)]*\)|（[^）]*）)$', teacher)
                            teacher_suffix = teacher_suffix_match.group(1) if teacher_suffix_match else ""
                            teacher_core = teacher[:-len(teacher_suffix)] if teacher_suffix else teacher
                            common_surnames = set(
                                "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦何吕施张曹严华金魏陶姜谢邹喻苏潘葛范彭鲁韦马苗方袁柳鲍史唐费薛雷贺倪汤殷罗毕郝邬安常乐于傅皮卞齐康伍余元顾孟平黄穆萧尹姚邵汪祁毛禹狄米明臧计成戴谈宋茅庞熊纪舒屈项祝董梁杜阮蓝闵席季麻强贾路江童颜郭梅盛林刁钟徐邱骆高夏蔡田樊胡凌霍虞万支柯管卢莫房裘缪干解应宗丁宣邓郁杭洪包诸左石崔吉钮龚程嵇邢裴陆荣翁荀羊惠甄曲封储靳松段富巫乌焦巴弓牧隗车侯全班仰秋仲伊宫宁仇栾暴甘厉戎祖武符刘景詹束龙叶幸司韶郜黎蓟薄印宿白怀蒲邰从鄂索赖卓蔺屠蒙池乔阴郁胥苍闻莘党翟谭贡劳逄姬申扶堵冉宰郦雍璩桑桂濮牛寿通边扈燕冀郏浦尚农温别庄晏柴瞿阎慕连茹习宦艾鱼容向古易慎戈廖庾终暨居衡步都耿满弘匡国文寇禄阙东欧殳沃利蔚越夔隆师巩厍聂晁勾敖融冷訾辛阚那简饶空曾毋沙乜养鞠须丰巢关蒯相查后荆红游竺权逯盖益桓公"
                            )
                            for size in (3, 2):
                                if len(teacher_core) <= size:
                                    continue
                                candidate = teacher_core[-size:]
                                if candidate and candidate[0] in common_surnames:
                                    name = teacher_core[:-size].strip()
                                    teacher = f"{candidate}{teacher_suffix}"
                                    break
                else:
                    # 如果没有 ()，尝试将最后一行作为教师名（前提是至少有两行）
                    if len(lines) >= 2:
                        teacher = lines[-1]
                        name = "\n".join(lines[:-1]).strip()
                    else:
                        # Single-line fallback: try "课程名 教师名" split by whitespace.
                        one_line = lines[0]
                        tokens = [t for t in one_line.split(' ') if t]
                        if len(tokens) >= 2:
                            teacher = tokens[-1]
                            name = " ".join(tokens[:-1]).strip()
                        else:
                            name = one_line
                            teacher = ""
            else:
                name = before_week
        else:
            # 如果没有周次标记之前的内容，设为未知
            name = "未知课程"
        
        # 🔥 提取地点
        # 地点通常在节次标记之后
        location_pattern = re.compile(r'【([^】]+)】([^【]*)')
        if teacher.endswith('()') or teacher.endswith('（）'):
            teacher = teacher[:-2]
        location_matches = location_pattern.findall(content)
        if not location_matches:
            location_matches = re.findall(r'\u3010([^\u3011]+)\u3011([^\u3010\u3011]*)', content)
        
        if location_matches:
            # 可能有多个匹配，取最后一个（通常是地点）
            campus, building = location_matches[-1]
            location = f"{campus}{building}".strip().replace('\n', '')
        else:
            # 尝试简单格式：节次标记后面的内容
            if section_match:
                after_section = content[section_match.end():].strip()
                # 提取第一个非空字符序列作为地点
                simple_location_match = re.match(r'^([A-Za-z\u4e00-\u9fa5]+\d+)', after_section)
                if simple_location_match:
                    location = simple_location_match.group(1)
            if not location:
                tail_location = re.search(r'([A-Za-z\u4e00-\u9fa5]+\d+)\s*$', content)
                if tail_location:
                    location = tail_location.group(1)
        
        # 清理课程名中的多余换行
        name = name.replace('\n', ' ').strip()
        if not name: name = "未知课程"

        # 🔥 B. 解析周次信息（支持离散周段）
        week_segments = []
        if bracket_week_match:
            # 解析 [周] 格式：2-9,11-17[周]
            week_str = bracket_week_match.group(1)
            print(f"✅ [Excel] 检测到 [周] 格式: {week_str}[周]")
            week_segments = self._parse_bracket_week_format(week_str)
        elif week_match:
            # 解析 {第...周} 格式
            start_week_str = week_match.group(1)
            end_week_str = week_match.group(2)
            week_type_str = week_match.group(3)
            
            start_week = int(start_week_str)
            end_week = int(end_week_str) if end_week_str else start_week
            week_type = WeekType.EVERY_WEEK
            if week_type_str == '单':
                week_type = WeekType.ODD_WEEK
            elif week_type_str == '双':
                week_type = WeekType.EVEN_WEEK
            
            week_segments = [{
                'start_week': start_week,
                'end_week': end_week,
                'week_type': week_type
            }]

        # 🔥 为每个周段生成一条 CourseDetail
        for seg in week_segments:
            print(f"✅ [Excel] 课程解析: '{name}' (教师: {teacher or '未知'}, 地点: {location or '未知'}, "
                  f"星期{day}, 第{actual_section}节, 持续{duration}节, "
                  f"周次: {seg['start_week']}-{seg['end_week']})")
            
            courses.append({
                'name': name,
                'teacher': teacher,
                'location': location,
                'section': actual_section,
                'duration': duration,
                'day': day,
                'start_week': seg['start_week'],
                'end_week': seg['end_week'],
                'week_type': seg['week_type']
            })
        
        return courses
    
    def _parse_bracket_week_format(self, week_str: str) -> List[dict]:
        """
        解析 [周] 格式的周次字符串
        
        例如: "2-9,11-17" → [{'start_week': 2, 'end_week': 9}, {'start_week': 11, 'end_week': 17}]
        
        Args:
            week_str: 周次字符串，如 "2-9,11-17" 或 "10"
            
        Returns:
            周段列表，每个元素包含 start_week, end_week, week_type
        """
        segments = []
        
        # 按逗号分段
        parts = week_str.split(',')
        
        for part in parts:
            part = part.strip()
            if '-' in part:
                # 范围格式：2-9
                start, end = part.split('-')
                segments.append({
                    'start_week': int(start.strip()),
                    'end_week': int(end.strip()),
                    'week_type': WeekType.EVERY_WEEK
                })
            else:
                # 单周格式：10
                week_num = int(part)
                segments.append({
                    'start_week': week_num,
                    'end_week': week_num,
                    'week_type': WeekType.EVERY_WEEK
                })
        
        print(f"📝 [Excel] 解析周段: {week_str} → {len(segments)} 个周段")
        return segments
    
    def _parse_standard_format(self, sheet: Worksheet) -> List[dict]:
        """
        解析标准格式的 Excel（简化格式）
        
        格式：课程名 教师 地点
        
        Args:
            sheet: openpyxl 工作表对象
            
        Returns:
            课程信息字典列表
        """
        import_beans = []
        
        # 获取所有行
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return import_beans
        
        # 查找表头行
        header_row_idx = 0
        weekday_col_map = {}
        
        for idx, row in enumerate(rows):
            for col_idx, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    for weekday_name, weekday_num in self.weekday_names.items():
                        if weekday_name in cell:
                            header_row_idx = idx
                            weekday_col_map[col_idx] = weekday_num
            
            if weekday_col_map:
                break
        
        # 如果没有找到表头，假设第一列是周一
        if not weekday_col_map:
            for i in range(1, 8):
                weekday_col_map[i] = i
        
        # 遍历数据行
        for row_idx in range(header_row_idx + 1, len(rows)):
            row = rows[row_idx]
            
            # 当前节次
            current_section = row_idx - header_row_idx
            
            # 遍历课程列
            for col_idx, cell_value in enumerate(row):
                # 跳过第一列（可能是节次列）
                if col_idx == 0:
                    continue
                
                # 跳过空单元格
                if not cell_value or str(cell_value).strip() == "":
                    continue
                
                # 获取星期
                day_of_week = weekday_col_map.get(col_idx, col_idx)
                
                # 解析单元格内容
                cell_text = str(cell_value).replace('\n', ' ').strip()
                beans = self._parse_simple_cell(cell_text, current_section, day_of_week)
                import_beans.extend(beans)
        
        return import_beans
    
    def _parse_simple_cell(self, content: str, section: int, day: int) -> List[dict]:
        """
        解析简单格式的单元格
        
        格式：课程名 教师 地点
        
        Args:
            content: 单元格文本
            section: 当前节次
            day: 星期几（1-7）
            
        Returns:
            课程信息字典列表
        """
        parts = content.split()
        
        if not parts:
            return []
        
        name = parts[0]
        teacher = parts[1] if len(parts) >= 2 else ""
        location = parts[2] if len(parts) >= 3 else ""
        
        return [{
            'name': name,
            'time_info': '{第1-16周',  # 默认周次
            'teacher': teacher,
            'location': location,
            'section': section,
            'day': day
        }]
    
    def _parse_sheet(self, sheet: Worksheet) -> List[dict]:
        """
        解析工作表
        
        预设模板：
        - 第一行：标题行（可能包含"周一"、"周二"等）
        - 第一列：节次信息（可能包含"第1-2节"等）
        - 其他单元格：课程信息
        
        Args:
            sheet: openpyxl 工作表对象
            
        Returns:
            课程信息字典列表
        """
        import_beans = []
        
        # 获取所有行
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return import_beans
        
        # 遍历所有单元格
        for row_idx, row in enumerate(rows):
            # 跳过第一行（标题行）
            if row_idx == 0:
                continue
            
            # 当前行的节次（从第1节开始）
            current_section = row_idx
            
            for col_idx, cell_value in enumerate(row):
                # 跳过第一列（节次列）
                if col_idx == 0:
                    continue
                
                # 跳过空单元格
                if not cell_value or str(cell_value).strip() == "":
                    continue
                
                # 将单元格内容转换为空格分隔的字符串
                cell_text = str(cell_value).replace('\n', ' ').strip()
                
                # 推断星期（第2列=周一，第3列=周二，...）
                day_of_week = col_idx
                
                # 解析单元格内容
                # 复用 HTML 导入器的解析逻辑
                beans = self._parse_cell_content(cell_text, current_section, day_of_week)
                import_beans.extend(beans)
        
        return import_beans
    
    def _parse_cell_content(self, content: str, section: int, day: int) -> List[dict]:
        """
        解析单元格内容
        
        单元格格式示例：
        "高等数学 {第1-16周 张老师 A101"
        "线性代数 {第1-16周(单) 李老师 B202"
        
        Args:
            content: 单元格文本
            section: 当前节次
            day: 星期几（1-7）
            
        Returns:
            课程信息字典列表
        """
        courses = []
        
        # 按空格分割
        parts = content.split()
        
        # 查找所有包含 { 的部分（周次信息）
        week_indices = []
        for i, part in enumerate(parts):
            if '{' in part:
                week_indices.append(i)
        
        # 如果没有找到周次信息，尝试简单格式
        if not week_indices:
            # 简单格式：课程名 教师 地点
            if len(parts) >= 1:
                name = parts[0]
                teacher = parts[1] if len(parts) >= 2 else ""
                location = parts[2] if len(parts) >= 3 else ""
                
                courses.append({
                    'name': name,
                    'time_info': f'周{self._get_day_name(day)}{{第1-16周',  # 默认周次
                    'teacher': teacher,
                    'location': location,
                    'section': section,
                    'day': day
                })
            return courses
        
        # 解析每个课程
        for idx, week_idx in enumerate(week_indices):
            # 课程名称在周次信息之前
            if week_idx == 0:
                continue
            
            name = parts[week_idx - 1]
            time_info = parts[week_idx]
            
            # 教师和地点在周次信息之后
            teacher = None
            location = None
            
            # 确定这个课程的结束位置
            if idx < len(week_indices) - 1:
                next_week_idx = week_indices[idx + 1]
                end_idx = next_week_idx - 1
            else:
                end_idx = len(parts)
            
            # 提取教师和地点
            remaining_parts = parts[week_idx + 1:end_idx]
            if len(remaining_parts) >= 1:
                teacher = remaining_parts[0]
            if len(remaining_parts) >= 2:
                location = remaining_parts[1]
            
            courses.append({
                'name': name,
                'time_info': time_info,
                'teacher': teacher or "",
                'location': location or "",
                'section': section,
                'day': day
            })
        
        return courses
    
    def _convert_to_courses(self, import_beans: List[dict]) -> Tuple[List[CourseBase], List[CourseDetail]]:
        """
        将导入的数据转换为 CourseBase 和 CourseDetail
        
        Args:
            import_beans: 导入的课程信息列表
            
        Returns:
            (CourseBase列表, CourseDetail列表)
        """
        course_bases = []
        course_details = []
        course_id_map = {}  # 课程名称 -> 课程ID 的映射
        
        for bean in import_beans:
            name = bean['name']
            
            # 检查课程是否已存在
            if name not in course_id_map:
                # 创建新的 CourseBase
                course_id = str(uuid.uuid4())
                color = self.color_manager.get_color_for_course(name)
                course_base = CourseBase(
                    name=name,
                    color=color,
                    course_id=course_id
                )
                course_bases.append(course_base)
                course_id_map[name] = course_id
            else:
                course_id = course_id_map[name]
            
            # 🔥 优先使用 bean 中已解析的周次信息
            if 'start_week' in bean and 'end_week' in bean:
                # 已经在 _parse_qiangzhi_cell 中解析过
                start_week = bean['start_week']
                end_week = bean['end_week']
                week_type = bean.get('week_type', WeekType.EVERY_WEEK)
                day = bean.get('day', 0)
            else:
                # 兼容旧格式：从 time_info 解析
                time_info = self._parse_time_info(bean.get('time_info', ''))
                start_week = time_info['start_week']
                end_week = time_info['end_week']
                week_type = time_info['week_type']
                day = time_info['day'] if time_info['day'] != 0 else bean.get('day', 0)
            
            # 🔥 使用解析出的duration
            duration = bean.get('duration', 2)
            
            # 创建 CourseDetail
            course_detail = CourseDetail(
                course_id=course_id,
                teacher=bean['teacher'],
                location=bean['location'],
                day_of_week=day,
                start_section=bean['section'],
                step=duration,
                start_week=start_week,
                end_week=end_week,
                week_type=week_type
            )
            course_details.append(course_detail)
        
        return course_bases, course_details
    
    def _parse_time_info(self, time_info: str) -> dict:
        """
        解析时间信息
        
        Args:
            time_info: 时间信息字符串，如 "{第1-16周" 或 "{第1-16周(单)"
            
        Returns:
            包含 day, step, start_week, end_week, week_type 的字典
        """
        result = {
            'day': 0,
            'step': 1,
            'start_week': 1,
            'end_week': 20,
            'week_type': WeekType.EVERY_WEEK
        }
        
        # 解析周次
        match = self.week_pattern.search(time_info)
        if match:
            start_week_str = match.group(1)
            end_week_str = match.group(2)
            week_type_str = match.group(3)
            
            result['start_week'] = int(start_week_str)
            
            if end_week_str:
                result['end_week'] = int(end_week_str)
            else:
                result['end_week'] = result['start_week']
            
            # 解析单双周
            if week_type_str == '单':
                result['week_type'] = WeekType.ODD_WEEK
            elif week_type_str == '双':
                result['week_type'] = WeekType.EVEN_WEEK
        
        return result
    
    def _get_day_name(self, day: int) -> str:
        """
        将数字转换为中文星期
        
        Args:
            day: 星期数字（1-7）
            
        Returns:
            中文星期，如 "一"
        """
        day_names = ["", "一", "二", "三", "四", "五", "六", "日"]
        if 1 <= day <= 7:
            return day_names[day]
        return ""
    
    def _extract_section_num(self, cell_value) -> Optional[int]:
        """
        从单元格中提取节次数字
        
        Args:
            cell_value: 单元格值，可能是 "1", "第1节", "1-2", 等
            
        Returns:
            节次数字，如果无法提取则返回None
        """
        if not cell_value:
            return None
        
        cell_str = str(cell_value).strip()
        
        # 尝试直接转换为数字
        try:
            return int(cell_str)
        except ValueError:
            pass
        
        # 尝试提取数字
        match = re.search(r'(\d+)', cell_str)
        if match:
            return int(match.group(1))
        
        return None
